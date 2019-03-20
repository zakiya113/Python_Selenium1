import openpyxl

#load workbook

wk= openpyxl.load_workbook("D:/TestSheet.xlsx")

print(wk.sheetnames)

print("Active Sheet is " + wk.active.title)

#Create Object of any sheet on which you want to work

sh = wk ['Sheet1']
print(sh.title)

#Fetch the Value
print(sh['A4'].value)
print(sh['B3'].value)
print(sh['C2'].value)

#Creating Cell Object

c1 = sh.cell(1,1)
print(c1.value)

#Find Rows having Data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
        c= sh.cell(i,j)
        print(c.value)

for r in sh['A1' : 'C4']:
    for c in r:
        print(c.value)